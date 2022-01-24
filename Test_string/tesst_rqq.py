
import openpyxl
import pandas as pd
from pandas import ExcelWriter
import re
import time
import os, inspect, sys
import datetime
from xlcalculator.tokenizer import col2num
import xlrd
from io import BytesIO, StringIO
import glob
import xlsxwriter
import shutil
import os
CurDir = os.path.dirname(os.path.abspath(inspect.getfile(inspect.currentframe())))
path_chrome = os.path.abspath(CurDir +"\\chromedriver.exe")
def round_int(x):
    last_dig = int(str(x)[-1])
    if last_dig >= 5:
        x += 10
    return (x/10) * 10
asdss =70183086.96
a =round(asdss)
print(a)
sdfsdf= 'dfgdfgdfgdfg'

# def path_input():
#     "Check folder input" 
#     path_folder_input = glob.glob(os.path.abspath(CurDir) + "\\*.xlsx")
#     # print(path_folder_input)
#     return path_folder_input
# lst_input =path_input()
# import openpyxl\


# path = r"D:\HuyNP\Huy/124.xlsx"

# wb = openpyxl.load_workbook(path, data_only= True)
# ws=wb["Sheet"]

# str_234 = ws.cell(row=11,column=4).value
# print(str_234)
# wb.save(path)

# def refresh_file(file):
#     xlapp = win32.DispatchEx("Excel.Application")
    # xlapp.Visible = False
    # xlapp.DisplayAlerts = False
    # xlapp.EnableEvents = False
#     path = os.path.abspath(file)
#     wb =  xlapp.Wordbooks.Open(path)
#     wb.RefreshAll()
#     xlapp.CalculateUntilAsyncqueriesDone()
#     wb.save()
#     xlapp.Quit()

# refresh_file(path)

path = r"D:\HuyNP\Huy\demo_pit_report_1641285136.xlsx"

# workbook = xlrd.open_workbook(path)
# worksheet = workbook.sheet_by_index(0)
# str_21 = worksheet.cell(22, 3).value
# print(str_21)
# excel_app = win32.DispatchEx("Excel.Application")

# excel_app.Visible = False
# excel_app.DisplayAlerts = False
# excel_app.EnableEvents = False
# xlBook = excel_app.Workbooks.Open(path)
# xlSheet = xlBook.Worksheets('Declaration (EN)')
# xlSheet.EnableCalculation = True
# print(xlSheet.cells(41,20).value)
# excel_app.Application.Quit()


# path = r"D:\HuyNP\Huy\demo_pit_report_1641281550.xlsx"
# df_input = pd.read_excel(path,sheet_name="Declaration (EN)", engine='openpyxl',dtype=object)

# print(df_input)


# with  pd.ExcelWriter(CurDir+"\\"+"gf.xlsx", mode=  'wb',
#                         engine='xlsxwriter', 
#                         options={'strings_to_urls': False,'strings_to_formulas': False}) as writer:
#     for key, values in df_input.items():
#         df_output = values.to_excel(writer,sheet_name=key, index = False)
#     writer.save() 


# import xlwings
# from openpyxl import load_workbook, Workbook
# book =  Workbook()

# import win32com.client as win32
# import os
# from pynput.keyboard import Key, Controller
# keyboard = Controller()
# path= r"D:\HuyNP\Huy\1asd4.xlsx"
# os.system('start excel.exe {}'.format(path))
# os.system('taskkill /T /IM EXCEL.exe')
# keyboard.press(Key.enter)
# keyboard.release(Key.enter)
# data = load_workbook(path,data_only=True)  
# ws = data["Sheet1"]
# str_234 = ws.cell(row=24,column=4).value
# print(str_234)
# excel_app = win32.DispatchEx("Excel.Application")

# excel_app.Visible = False
# excel_app.DisplayAlerts = False
# excel_app.EnableEvents = False
# xlBook = excel_app.Workbooks.Open(path)

# xlBook.save()
# excel_app.Application.Quit()


# from openpyxl import Workbook
# path = r"D:\HuyNP\Huy/local_filename.xlsx"

def load_values(path):

    import formulas
    wb = openpyxl.load_workbook(path)
    ws = wb["'Declaration (EN)'"]
    int_value = ws.cell(41, 20).value
    print(int_value)
   

  
    


   
 


load_values(path)


# wb = openpyxl.load_workbook(path)
# Sheet_name =  wb.active

# str_234 = Sheet_name.cell(row=23,column=4).value
# print(str_234)





# import xlwings as xw
# app = xw.App(visible=False)
# readsheet = xw.Book(path).sheets['Declaration (EN)']
# df = pd.DataFrame(readsheet.range('A1', 'Z99').value)
# print(df)
# wb = openpyxl.load_workbook(path,data_only=True)
# ws = wb["Declaration (EN)"]
# str_123= ws.cell(row=46,column=20).value
# print(str_123)


# df_input["Unnamed: 6"][100] = "123"
# df_output = df_input
# with ExcelWriter("D:\HuyNP\Huy"+"\\"+"demo_pit_report_1634630031 (2).xlsx", engine= "openpyxl")as writer:
                    
#     df_output.to_excel(writer,sheet_name="Declaration (EN)",index=False)
#     writer.save()
# print(df_input)

path = r"D:\HuyNP\Huy\demo_pit_report_1641285136.xlsx"

def data_input():

    df_input = pd.read_excel(path,sheet_name="Declaration (EN)", engine='openpyxl',dtype=object)
    # print(df_input[:15])
    for idx, row in df_input.iterrows():
        
        if str(row).find('[01]') and str(row).find("Quarter") != -1:

            str_Quykekhai = (re.findall('(?<=\(Quarter\)\s)\d{1}',df_input.iloc[idx,1]))[0]
            str_Namkekhai = (re.findall('(?<=\(Year\)\s)\d{4}',df_input.iloc[idx,1]))[0]
            str_Thangkekhai = ""

        if str(row).find('[01]') and str(row).find("Month") != -1:
            str_Namkekhai = (re.findall('(?<=\(Year\)\s)\d{4}',df_input.iloc[idx,1]))[0]
            str_Thangkekhai = (re.findall('(?<=\(Month\)\s)\d{1,2}',df_input.iloc[idx,1]))[0]
            str_Quykekhai = ""
        if str(row).find('[02]') != -1:
            # print(row)

            if str(row).find("Lần đầu (First time):     [ X ]") != -1:
                # print(row)
                print("To khai chinh thuc")      
            else :
                # print(row)
                print("To khai bo sung")
        if str(row).find('[05]') != -1:
            str_MST = str(df_input.iloc[idx,3])+ str(df_input.iloc[idx,4])+str(df_input.iloc[idx,5])+str(df_input.iloc[idx,6])+str(df_input.iloc[idx,7])+str(df_input.iloc[idx,8])+str(df_input.iloc[idx,9])+str(df_input.iloc[idx,10])+str(df_input.iloc[idx,11])+str(df_input.iloc[idx,12])+str(df_input.iloc[idx,13]).strip()+str(df_input.iloc[idx,14])+str(df_input.iloc[idx,15])+str(df_input.iloc[idx,16])
            str_MST = str_MST.split("nan")[0]
        
        if str(row).find('[21]') != -1:
            # str_CT21 = get_Value(path, row= idx, column=19)
            str_CT21 = str(df_input.iloc[idx,19])

        if str(row).find('[22]') != -1:
            # str_CT22 = get_Value(path, row= idx, column=19)
            str_CT22 = str(df_input.iloc[idx,19])

        # if str(row).find('[23]') != -1:
        #     str_CT23 = get_Value(r"D:\HuyNP\Huy\local_filename.xlsx", row= idx, column=19)
        
        if str(row).find('[24]') != -1:
            # str_CT24 = get_Value(path, row= idx, column=19)
            str_CT24 = str(df_input.iloc[idx,19])

        if str(row).find('[25]') != -1:
            str_CT25 = str(df_input.iloc[idx,19])

        if str(row).find('[26]') != -1:
            str_26 = str(df_input.iloc[idx,19])
            # str_CT26 = get_Value(path, row= idx, column=19)
            # print(str_26)

        if str(row).find('[27]') != -1:
            str_CT27 = str(df_input.iloc[idx,19])
        if str(row).find('[28]') != -1:
            str_CT28 = str(df_input.iloc[idx,19])
            
        if str(row).find('[29]') != -1:
            # str_CT29 = get_Value(path, row= idx, column=19)
            str_CT29 = str(df_input.iloc[idx,19])
            
        if str(row).find('[30]') != -1:
            str_CT30 = str(df_input.iloc[idx,19])

        if str(row).find('[31]') != -1:
            str_CT31 = str_CT28
            # print(str_CT31)

        if str(row).find('[32]') != -1:
            # str_CT32 = get_Value(path, row= idx, column=19)
            str_CT32 = str(df_input.iloc[idx,19])

        if str(row).find('[33]') != -1:
            # str_CT33 = get_Value(path, row= idx, column=19)
            str_CT33 = str(df_input.iloc[idx,19])

        if str(row).find('[34]') != -1:
            str_CT34 = str(df_input.iloc[idx,19])  
'[02] Lần đầu (First time):     [    ]'          
'[02] Lần đầu (First time):     [ X ]'
'[03] Bổ sung lần thứ (Suplementary):     [ X ]'
# data_input()
def get_Value(filename = "", row=None ,column= None):
    # row = idx+1,  cols = 20
    import openpyxl
    from xlcalculator import ModelCompiler
    from xlcalculator import Model
    from xlcalculator import Evaluator
    if row != None and column != None:
        work_book = openpyxl.load_workbook(filename)
        work_sheet = work_book["Declaration (EN)"]

        int_value = work_sheet.cell(row+2, column+1).value
        str_value = int_value.replace("=","")
        compiler = ModelCompiler()
        new_model = compiler.read_and_parse_archive(filename, ignore_sheets="1")
        evaluator = Evaluator(new_model)
        val1 = evaluator.evaluate(str_value)
        return str(val1)
    else:
        val1 = ""
        return val1



def demo_data():

    # a =os.path.abspath("Huy\\"+'local_filename.xlsx')
    path = r"D:\HuyNP\Huy\demo_pit_report_1641281550.xlsx"
    import win32com.client as win32
    import os
    import win32com.client
 
    xlApp =win32com.client.dynamic.Dispatch("Excel.Application")
    # xlApp.DisplayAlerts = False
    xlwb = xlApp.Workbooks.Open(path, True, False, None)
    # xlApp.DisplayAlerts = True
    xlSheet = xlwb.Worksheets('Declaration (EN)')
    print(int(xlSheet.cells(46,20).value))
    xlwb.Close(True)
    del xlApp

# demo_data()



import requests
def DownloadFile(url):
    local_filename = url.split('?')[0].split('/')[-1]
    r = requests.get(url)
    f = open(r'D:\HuyNP\Huy\local_filename.xlsx', 'wb')
    for chunk in r.iter_content(chunk_size=512 * 1024): 
        if chunk: # filter out keep-alive new chunks
            f.write(chunk)
    f.close()
     
# DownloadFile(r"https://media-dev.vina-payroll.com/media/ReportPit/1067/vs026_pit_report_1640321957.xlsx?X-Amz-Content-Sha256=UNSIGNED-PAYLOAD&X-Amz-Algorithm=AWS4-HMAC-SHA256&X-Amz-Credential=gDJyyFZPgWhLWP9suF1m%2F20220120%2Fus-east-1%2Fs3%2Faws4_request&X-Amz-Date=20220120T071401Z&X-Amz-SignedHeaders=host&X-Amz-Expires=300&X-Amz-Signature=bfde5812a58c2e3082a20548239beeba80a698525560e65f377f7d5df5ac82a0")

# import io
# import chardet
# import codecs


# raw = open(r'D:\HuyNP\Huy\local_filename.txt', 'rb').read(32)

# if raw.startswith(codecs.BOM_UTF8):
#     encoding = 'utf-8-sig'
# else:
#     result = chardet.detect(raw)
#     encoding = result['encoding']

# infile = open(r'D:\HuyNP\Huy\local_filename.txt',encoding=encoding,strict=False)
# data = infile.read()
# infile.close()

# print(data)




# def get_Report(urls):
#     from io import BytesIO, StringIO

#     name_split = urls.split('?')[0].split('/')[-1]
   
#     payload="{\"query\":\"\",\"variables\":{}}"
#     headers = {
#     'Authorization': 'Basic bGVhZGVyX3Rlc3Q6MTIzNDU2',
#     'Content-Type': 'application/json'
#     }

#     response = requests.request("GET", urls, headers=headers, data=payload)
#     BytesIO = pd.io.common.BytesIO
#     df_result_hd = pd.read_excel(BytesIO(response.content), sheet_name=0, dtype=object)

#     with ExcelWriter("D:\HuyNP\Huy"+"\\"+name_split, engine= "xlsxwriter",mode="wb",
#                         options={'strings_to_urls': False,'strings_to_formulas': False})as writer:
#         df_result_hd.to_excel(writer,sheet_name="Declaration (EN)",index=False)
#         writer.save()

    # df_result=pd.read_excel(io.StringIO(response.decode("utf-8")))
    
  
    

# get_Report('https://media-dev.vina-payroll.com/media/ReportPit/869/demo_pit_report_1634630031.xlsx?X-Amz-Content-Sha256=UNSIGNED-PAYLOAD&X-Amz-Algorithm=AWS4-HMAC-SHA256&X-Amz-Credential=gDJyyFZPgWhLWP9suF1m%2F20211215%2Fus-east-1%2Fs3%2Faws4_request&X-Amz-Date=20211215T015121Z&X-Amz-SignedHeaders=host&X-Amz-Expires=300&X-Amz-Signature=89483098105bdf9582302cda862f59edf13df44be968ef9510cbe4c42f872917')


